##import all the libraries
import os
import openpyxl
import pandas as pd
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string
import tkinter as tk
from tkinter.ttk import *
from tkinter.filedialog import askopenfile
import openpyxl


    
root = tk.Tk()
root.title("SortApp")
root.geometry('400x400')
  


def open_file():
    browse_text.set("loading...")
    file = askopenfile(mode ='rb', filetypes =[('Excel Files', '*.xlsx *.xlsm *.sxc *.ods *.csv *.tsv')])
    if file :        
        wb = openpyxl.load_workbook(file)
        ws = wb['Master']
        
        #text box
        text_box = tk.Text(root, height=3, width=10, padx=15, pady=15)
        text_box.insert(1.0, ws)
        text_box.tag_configure("center", justify="center")
        text_box.tag_add("center", 1.0, "end")
        text_box.grid(column=1, row=3)

        browse_text.set("Import the file")
        print('file Imported')
        

        ##Create a new Sheet "Monday"
        Mon = wb.create_sheet(title = 'Monday')
        
        #concvert the data into list
        code =[]
        desc1=[]
        desc2=[]
        row_count= ws.max_row
        rows = ws.iter_rows(min_row=1, max_row=row_count-1, min_col=1, max_col=3)
        for a,b,c in rows:
            code.append(a.value)
            desc1.append(b.value)
            desc2.append(c.value)

        df =pd.DataFrame([code,desc1,desc2])
        df1= df.transpose()   
        new_header = df1.iloc[0] 
        #df1 = df1[1:] 
        df1.columns = new_header


        ## Sort and filter the code for monday
        MON_10 = df1[df1['Description 2'].str.contains('M') & df1['Description 2'].str.contains('1000')]
        MON_11 = df1[df1['Description 2'].str.contains('M') & df1['Description 2'].str.contains('1100')]
        MON_12 = df1[df1['Description 2'].str.contains('M') & df1['Description 2'].str.contains('1200')]
        MON_13 = df1[df1['Description 2'].str.contains('M') & df1['Description 2'].str.contains('1300')]
        MON_14 = df1[df1['Description 2'].str.contains('M') & df1['Description 2'].str.contains('1400')]
        MON_143 = df1[df1['Description 2'].str.contains('M') & df1['Description 2'].str.contains('1430')]
        MON_15= df1[df1['Description 2'].str.contains('M') & df1['Description 2'].str.contains('1500')]
        MON_16 = df1[df1['Description 2'].str.contains('M') & df1['Description 2'].str.contains('1600')]

        # Change the dataframe to list  
        mon10 =list(MON_10['Code'])
        mon11= list(MON_11['Code'])
        mon12=list(MON_12['Code'])
        mon13 =list(MON_13['Code'])
        mon14 = list(MON_14['Code'])
        mon143= list(MON_143['Code'])
        mon15 =list(MON_15['Code'])
        mon16=list(MON_16['Code'])    



        #Insert the title head
        Mon.cell(column=1, row =2, value='MONDAY')
        Mon.cell(column=1, row =3, value='DATE')
        Mon.cell(column=1, row =4, value='TIME')
        Time =['10:00AM','11:00AM','12:00NOON','13:00PM','14:00PM','14:30PM','15:00PM','16:00PM']
        Code= ['Code','Code','Code','Code','Code','Code','Code','Code']
        for i in range(len(Time)):
            Mon.cell(column=i+1, row =5, value=Time[i])
            Mon.cell(column=i+1, row =6, value=Code[i])


        ## Insert the stop code into the excel sheet  
        for i in range(len(mon10)):
            Mon.cell(column=1, row =i+7, value=mon10[i])

        for i in range(len(mon11)):
            Mon.cell(column=2, row =i+7, value=mon11[i])

        for i in range(len(mon12)):
            Mon.cell(column=3, row =i+7, value=mon12[i])

        for i in range(len(mon13)):
            Mon.cell(column=4, row =i+7, value=mon13[i])

        for i in range(len(mon14)):
            Mon.cell(column=5, row =i+7, value=mon14[i])

        for i in range(len(mon143)):
            Mon.cell(column=6, row =i+7, value=mon143[i])

        for i in range(len(mon15)):
            Mon.cell(column=7, row =i+7, value=mon15[i])

        for i in range(len(mon16)):
            Mon.cell(column=8, row =i+7, value=mon16[i])




        ##Formating the cell
        font = Font(name='Times New Roman',
                        size=20,
                        bold=True,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False,
                        color='FF000000')
        fill_cell = PatternFill(fill_type=None,
                        start_color='FFFFFFFF',
                         end_color='FF000000')
        thin_border = Border(left=Side(border_style='dashed', color='FF000000'),
                         right=Side(border_style='dashed', color='FF000000'),
                         top=Side(border_style='thin',  color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000'))
        thick_border = Border(left=Side(border_style='thin', color='FF000000'),
                         right=Side(border_style='thin', color='FF000000'),
                         top=Side(border_style='thin',  color='FF000000'),
                         bottom=Side(border_style='thin', color='FF000000'))
        double_border = Border(left=Side(border_style='double', color='FF000000'),
                         right=Side(border_style='double', color='FF000000'),
                         top=Side(border_style='double',  color='FF000000'),
                         bottom=Side(border_style='double', color='FF000000'))
        alignment =      Alignment(horizontal='center',
                            vertical='center',
                            text_rotation=0,
                            wrap_text=False,
                            shrink_to_fit=False,
                            indent=0)

        ##Formating the cells
        row_num=Mon.max_row
        col_num=Mon.max_column
        row_loc=0
        col_loc=0
        for i in range(row_loc,row_loc+row_num):
            for j in range(col_loc,col_num+col_loc):
                Mon.cell(row=i+1,column=j+1).border=thick_border
                Mon.cell(row=i+1,column=j+1).fill=fill_cell



        for i in range(row_num):
            Mon.row_dimensions[i].height = 23

        for col in range(1,col_num):
            Mon.column_dimensions[get_column_letter(col)].width = 15

        Mon.cell(column=1, row =1, value="Sort into different time")
        Mon.merge_cells('A1:H1')


        Mon['A1'].font = font
        Mon.cell(row=1,column=1).alignment=alignment
        print("Monday is done")
        

        ##Create a new Sheet "Tuesday"
        Tue = wb.create_sheet(title = 'Tuesday')
        
        

        ## Sort and filter the code for Tuesday
        TUE_10 = df1[df1['Description 2'].str.contains('T') & df1['Description 2'].str.contains('1000')]
        TUE_11 = df1[df1['Description 2'].str.contains('T') & df1['Description 2'].str.contains('1100')]
        TUE_12 = df1[df1['Description 2'].str.contains('T') & df1['Description 2'].str.contains('1200')]
        TUE_13 = df1[df1['Description 2'].str.contains('T') & df1['Description 2'].str.contains('1300')]
        TUE_14 = df1[df1['Description 2'].str.contains('T') & df1['Description 2'].str.contains('1400')]
        TUE_143 = df1[df1['Description 2'].str.contains('T') & df1['Description 2'].str.contains('1430')]
        TUE_15= df1[df1['Description 2'].str.contains('T') & df1['Description 2'].str.contains('1500')]
        TUE_16 = df1[df1['Description 2'].str.contains('T') & df1['Description 2'].str.contains('1600')]

        # Change the dataframe to list  
        tue10 =list(TUE_10['Code'])
        tue11= list(TUE_11['Code'])
        tue12=list(TUE_12['Code'])
        tue13 =list(TUE_13['Code'])
        tue14 = list(TUE_14['Code'])
        tue143= list(TUE_143['Code'])
        tue15 =list(TUE_15['Code'])
        tue16=list(TUE_16['Code'])    



        #Insert the title head
        Tue.cell(column=1, row =2, value='Tuesday')
        Tue.cell(column=1, row =3, value='DATE')
        Tue.cell(column=1, row =4, value='TIME')
        Time =['10:00AM','11:00AM','12:00NOON','13:00PM','14:00PM','14:30PM','15:00PM','16:00PM']
        Code= ['Code','Code','Code','Code','Code','Code','Code','Code']
        for i in range(len(Time)):
            Tue.cell(column=i+1, row =5, value=Time[i])
            Tue.cell(column=i+1, row =6, value=Code[i])


        ## Insert the stop code into the excel sheet  
        for i in range(len(tue10)):
            Tue.cell(column=1, row =i+7, value=tue10[i])

        for i in range(len(tue11)):
            Tue.cell(column=2, row =i+7, value=tue11[i])

        for i in range(len(tue12)):
            Tue.cell(column=3, row =i+7, value=tue12[i])

        for i in range(len(tue13)):
            Tue.cell(column=4, row =i+7, value=tue13[i])

        for i in range(len(tue14)):
            Tue.cell(column=5, row =i+7, value=tue14[i])

        for i in range(len(tue143)):
            Tue.cell(column=6, row =i+7, value=tue143[i])

        for i in range(len(tue15)):
            Tue.cell(column=7, row =i+7, value=tue15[i])

        for i in range(len(tue16)):
            Tue.cell(column=8, row =i+7, value=tue16[i])




       

        ##Formating the cells
        row_num=Tue.max_row
        col_num=Tue.max_column
        row_loc=0
        col_loc=0
        for i in range(row_loc,row_loc+row_num):
            for j in range(col_loc,col_num+col_loc):
                Tue.cell(row=i+1,column=j+1).border=thick_border
                Tue.cell(row=i+1,column=j+1).fill=fill_cell



        for i in range(row_num):
            Tue.row_dimensions[i].height = 23

        for col in range(1,col_num):
            Tue.column_dimensions[get_column_letter(col)].width = 15

        Tue.cell(column=1, row =1, value="Sort into different time")
        Tue.merge_cells('A1:H1')


        Tue['A1'].font = font
        Tue.cell(row=1,column=1).alignment=alignment
       
        #wb.save('Sorted.xlsx')
        print('Tuesday is Done')


        
        ##Create a new Sheet "Wednesday"
        Wed = wb.create_sheet(title = 'Wednesday')
        

        ## Sort and filter the code for Wednesday
        WED_10 = df1[df1['Description 2'].str.contains('W') & df1['Description 2'].str.contains('1000')]
        WED_11 = df1[df1['Description 2'].str.contains('W') & df1['Description 2'].str.contains('1100')]
        WED_12 = df1[df1['Description 2'].str.contains('W') & df1['Description 2'].str.contains('1200')]
        WED_13 = df1[df1['Description 2'].str.contains('W') & df1['Description 2'].str.contains('1300')]
        WED_14 = df1[df1['Description 2'].str.contains('W') & df1['Description 2'].str.contains('1400')]
        WED_143 = df1[df1['Description 2'].str.contains('W') & df1['Description 2'].str.contains('1430')]
        WED_15= df1[df1['Description 2'].str.contains('W') & df1['Description 2'].str.contains('1500')]
        WED_16 = df1[df1['Description 2'].str.contains('W') & df1['Description 2'].str.contains('1600')]

        # Change the dataframe to list  
        wed10 =list(WED_10['Code'])
        wed11= list(WED_11['Code'])
        wed12=list(WED_12['Code'])
        wed13 =list(WED_13['Code'])
        wed14 = list(WED_14['Code'])
        wed143= list(WED_143['Code'])
        wed15 =list(WED_15['Code'])
        wed16=list(WED_16['Code'])    



        #Insert the title head
        Wed.cell(column=1, row =2, value='Wednesday')
        Wed.cell(column=1, row =3, value='DATE')
        Wed.cell(column=1, row =4, value='TIME')
        Time =['10:00AM','11:00AM','12:00NOON','13:00PM','14:00PM','14:30PM','15:00PM','16:00PM']
        Code= ['Code','Code','Code','Code','Code','Code','Code','Code']
        for i in range(len(Time)):
            Wed.cell(column=i+1, row =5, value=Time[i])
            Wed.cell(column=i+1, row =6, value=Code[i])


        ## Insert the stop code into the excel sheet  
        for i in range(len(wed10)):
            Wed.cell(column=1, row =i+7, value=wed10[i])

        for i in range(len(wed11)):
            Wed.cell(column=2, row =i+7, value=wed11[i])

        for i in range(len(wed12)):
            Wed.cell(column=3, row =i+7, value=wed12[i])

        for i in range(len(tue13)):
            Wed.cell(column=4, row =i+7, value=wed13[i])

        for i in range(len(wed14)):
            Wed.cell(column=5, row =i+7, value=wed14[i])

        for i in range(len(wed143)):
            Wed.cell(column=6, row =i+7, value=wed143[i])

        for i in range(len(wed15)):
            Wed.cell(column=7, row =i+7, value=wed15[i])

        for i in range(len(wed16)):
            Wed.cell(column=8, row =i+7, value=wed16[i])




       

        ##Formating the cells
        row_num=Wed.max_row
        col_num=Wed.max_column
        row_loc=0
        col_loc=0
        for i in range(row_loc,row_loc+row_num):
            for j in range(col_loc,col_num+col_loc):
                Wed.cell(row=i+1,column=j+1).border=thick_border
                Wed.cell(row=i+1,column=j+1).fill=fill_cell



        for i in range(row_num):
            Wed.row_dimensions[i].height = 23

        for col in range(1,col_num):
            Wed.column_dimensions[get_column_letter(col)].width = 15

        Wed.cell(column=1, row =1, value="Sort into different time")
        Wed.merge_cells('A1:H1')


        Wed['A1'].font = font
        Wed.cell(row=1,column=1).alignment=alignment
       
        #wb.save('Sorted.xlsx')
        print('Wednesday is Done')

        ##Thursday
        ##Create a new Sheet "Thursday"
        Thur = wb.create_sheet(title = 'Thursday')
       


        ## Sort and filter the code for Tuesday
        THUR_10 = df1[df1['Description 2'].str.contains('R') & df1['Description 2'].str.contains('1000')]
        THUR_11 = df1[df1['Description 2'].str.contains('R') & df1['Description 2'].str.contains('1100')]
        THUR_12 = df1[df1['Description 2'].str.contains('R') & df1['Description 2'].str.contains('1200')]
        THUR_13 = df1[df1['Description 2'].str.contains('R') & df1['Description 2'].str.contains('1300')]
        THUR_14 = df1[df1['Description 2'].str.contains('R') & df1['Description 2'].str.contains('1400')]
        THUR_143 = df1[df1['Description 2'].str.contains('R') & df1['Description 2'].str.contains('1430')]
        THUR_15= df1[df1['Description 2'].str.contains('R') & df1['Description 2'].str.contains('1500')]
        THUR_16 = df1[df1['Description 2'].str.contains('R') & df1['Description 2'].str.contains('1600')]

        # Change the dataframe to list  
        thur10 =list(THUR_10['Code'])
        thur11= list(THUR_11['Code'])
        thur12=list(THUR_12['Code'])
        thur13 =list(THUR_13['Code'])
        thur14 = list(THUR_14['Code'])
        thur143= list(THUR_143['Code'])
        thur15 =list(THUR_15['Code'])
        thur16=list(THUR_16['Code'])    



        #Insert the title head
        Thur.cell(column=1, row =2, value='Thursday')
        Thur.cell(column=1, row =3, value='DATE')
        Thur.cell(column=1, row =4, value='TIME')
        Time =['10:00AM','11:00AM','12:00NOON','13:00PM','14:00PM','14:30PM','15:00PM','16:00PM']
        Code= ['Code','Code','Code','Code','Code','Code','Code','Code']
        for i in range(len(Time)):
            Thur.cell(column=i+1, row =5, value=Time[i])
            Thur.cell(column=i+1, row =6, value=Code[i])


        ## Insert the stop code into the excel sheet  
        for i in range(len(thur10)):
            Thur.cell(column=1, row =i+7, value=thur10[i])

        for i in range(len(thur11)):
            Thur.cell(column=2, row =i+7, value=thur11[i])

        for i in range(len(thur12)):
            Thur.cell(column=3, row =i+7, value=thur12[i])

        for i in range(len(thur13)):
            Thur.cell(column=4, row =i+7, value=thur13[i])

        for i in range(len(thur14)):
            Thur.cell(column=5, row =i+7, value=thur14[i])

        for i in range(len(thur143)):
            Thur.cell(column=6, row =i+7, value=thur143[i])

        for i in range(len(thur15)):
            Thur.cell(column=7, row =i+7, value=thur15[i])

        for i in range(len(thur16)):
            Thur.cell(column=8, row =i+7, value=thur16[i])




       

        ##Formating the cells
        row_num=Thur.max_row
        col_num=Thur.max_column
        row_loc=0
        col_loc=0
        for i in range(row_loc,row_loc+row_num):
            for j in range(col_loc,col_num+col_loc):
                Thur.cell(row=i+1,column=j+1).border=thick_border
                Thur.cell(row=i+1,column=j+1).fill=fill_cell



        for i in range(row_num):
            Thur.row_dimensions[i].height = 23

        for col in range(1,col_num):
            Thur.column_dimensions[get_column_letter(col)].width = 15

        Thur.cell(column=1, row =1, value="Sort into different time")
        Thur.merge_cells('A1:H1')


        Thur['A1'].font = font
        Thur.cell(row=1,column=1).alignment=alignment
       
        
        print('Thursday is Done')



        
       
        wb.save('Sorted.xlsx')
        print('File saved')


       
  
#browse button
browse_text = tk.StringVar()
browse_btn = tk.Button(root, textvariable=browse_text, command=lambda:open_file(), font="Raleway", bg="#20bebe", fg="white", height=2, width=15)
browse_text.set("Browse")
browse_btn.grid(column=1, row=2)


# # # #Load the file

# monday = tk.Button(root, text='load the file',command =lambda:Monday(wb), font="Raleway", bg="#20bebe", fg="white", height=2, width=15)

# monday.grid(column=3, row=3)
  
root.mainloop()    
